import openpyxl

# Create a test Excel file with realistic data
# This will help identify if the issue is with the data format or the validation logic

def create_realistic_test_excel():
    """Create a test Excel file with realistic member data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Members"
    
    # Headers matching the backend exactly
    headers = [
        'First Name', 'Last Name', 'Gender', 'Date of Birth', 
        'Phone Number', 'Email', 'Address', 'Password',
        'Emergency Contact Name', 'Emergency Contact Phone',
        'Membership Plan Name', 'Status', 'Start Date', 'End Date', 'Medical Notes'
    ]
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Add test data with realistic values
    # NOTE: You need to replace "Gold Plan" with an actual plan name from your database
    test_data = [
        ['John', 'Smith', 'Male', '15-01-1990', '9876543210', 'john.smith@test.com', '123 Main Street, City', 'TestPass123', 'Jane Smith', '9876543211', 'Gold Plan', 'Active', '10-07-2026', '10-07-2027', 'No allergies'],
        ['Sarah', 'Johnson', 'Female', '20-05-1992', '9876543212', 'sarah.j@test.com', '456 Oak Avenue, Town', 'TestPass123', 'Mike Johnson', '9876543213', 'Silver Plan', 'Active', '10-07-2026', '10-07-2027', 'None'],
        ['Michael', 'Williams', 'Male', '10-03-1988', '9876543214', 'michael.w@test.com', '789 Pine Road, Village', 'TestPass123', 'Lisa Williams', '9876543215', 'Gold Plan', 'Active', '10-07-2026', '10-07-2027', 'Asthma']
    ]
    
    for row_num, row_data in enumerate(test_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    wb.save('test_realistic_bulk_upload.xlsx')
    print("Realistic test Excel file created: test_realistic_bulk_upload.xlsx")
    print("\nIMPORTANT: Before uploading, check your database for existing membership plans.")
    print("The template uses: 'Gold Plan', 'Silver Plan'")
    print("If your database has different plan names (e.g., 'Gold', 'Silver'), update the Excel file accordingly.")
    print("\nTo check existing plans in your database:")
    print("SELECT plan_name FROM membership_plans WHERE gym_id = YOUR_GYM_ID;")

if __name__ == '__main__':
    create_realistic_test_excel()
