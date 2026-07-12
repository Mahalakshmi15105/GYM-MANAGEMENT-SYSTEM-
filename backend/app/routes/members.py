from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt
from sqlalchemy import or_
from app.extensions import db, bcrypt
from app.models import Member, MembershipPlan
from app.activity_logging import ActivityLogger
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
import io
import csv

members_bp = Blueprint('members', __name__)

def get_current_gym_id():
    """Extract gym_id from JWT token for multi-tenant isolation"""
    claims = get_jwt()
    return claims.get('gym_id')

@members_bp.route('', methods=['GET'])
@jwt_required()
def list_members():
    """List all members for the current gym (multi-tenant filtered)"""
    gym_id = get_current_gym_id()
    if not gym_id:
        return jsonify({'error': 'Gym ID not found in token'}), 400
    
    # Get query parameters for filtering and search
    status_filter = request.args.get('status')
    search_query = request.args.get('q', '').strip()
    
    # Base query with multi-tenant isolation
    query = Member.query.filter_by(gym_id=gym_id)
    
    # Apply status filter
    if status_filter and status_filter != 'All':
        query = query.filter(Member.status == status_filter)
    
    # Apply search filter
    if search_query:
        search_pattern = f"%{search_query}%"
        query = query.filter(
            or_(
                Member.first_name.ilike(search_pattern),
                Member.last_name.ilike(search_pattern),
                Member.email.ilike(search_pattern),
                Member.phone.ilike(search_pattern),
                Member.member_id.ilike(search_pattern)
            )
        )
    
    # Order by created date (newest first) and limit for safety
    members = query.order_by(Member.created_at.desc()).limit(100).all()
    
    # Log the view operation
    ActivityLogger.log_view('member', view_type='list', gym_id=gym_id)
    
    return jsonify({
        'members': [member.to_dict() for member in members]
    }), 200

@members_bp.route('', methods=['POST'])
@jwt_required()
def create_member():
    """Create a new member for the current gym"""
    gym_id = get_current_gym_id()
    if not gym_id:
        return jsonify({'error': 'Gym ID not found in token'}), 400
    
    data = request.get_json() or {}
    
    # Required fields validation
    if not data.get('first_name') or not data.get('phone'):
        return jsonify({'error': 'First name and phone are required'}), 400
    
    # Check unique constraints within gym
    if data.get('email'):
        existing_email = Member.query.filter_by(gym_id=gym_id, email=data['email']).first()
        if existing_email:
            return jsonify({'error': 'Email already exists for a member in this gym'}), 400
        
    existing_phone = Member.query.filter_by(gym_id=gym_id, phone=data['phone']).first()
    if existing_phone:
        return jsonify({'error': 'Phone number already exists for a member in this gym'}), 400
    
    try:
        # Parse dates
        start_date = None
        end_date = None
        dob = None
        
        if data.get('membership_start_date'):
            start_date = datetime.strptime(data['membership_start_date'], '%Y-%m-%d').date()
        if data.get('membership_end_date'):
            end_date = datetime.strptime(data['membership_end_date'], '%Y-%m-%d').date()
        if data.get('date_of_birth'):
            dob = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()
        
        # Generate unique member_id
        import uuid
        member_id = f"MEM{str(uuid.uuid4())[:8].upper()}"
        
        # Ensure member_id is unique within gym
        while Member.query.filter_by(gym_id=gym_id, member_id=member_id).first():
            member_id = f"MEM{str(uuid.uuid4())[:8].upper()}"
        
        # Hash password if provided
        password_hash = None
        if data.get('password'):
            password_hash = bcrypt.generate_password_hash(data['password']).decode('utf-8')
        
        # Create new member
        new_member = Member(
            gym_id=gym_id,
            member_id=member_id,
            first_name=data['first_name'],
            last_name=data.get('last_name', ''),
            gender=data.get('gender', ''),
            date_of_birth=dob,
            phone=data['phone'],
            email=data.get('email', ''),
            password_hash=password_hash,
            address=data.get('address', ''),
            emergency_contact_name=data.get('emergency_contact_name', ''),
            emergency_contact_phone=data.get('emergency_contact_phone', ''),
            medical_notes=data.get('medical_notes', ''),
            membership_plan_name=data.get('membership_plan_name', ''),
            membership_start_date=start_date,
            membership_end_date=end_date,
            status=data.get('status', 'Active'),
            workout_duration_minutes=int(data.get('workout_duration_minutes', 120))
        )
        
        db.session.add(new_member)
        db.session.commit()
        
        # Log the create operation
        member_name = f"{new_member.first_name} {new_member.last_name}".strip()
        ActivityLogger.log_create(
            'member', 
            new_member.id, 
            entity_name=member_name,
            gym_id=gym_id,
            extra_data={'phone': new_member.phone, 'email': new_member.email}
        )
        
        return jsonify({
            'message': 'Member created successfully',
            'member': new_member.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to create member: {str(e)}'}), 500


@members_bp.route('/<int:member_id>', methods=['GET'])
@jwt_required()
def get_member(member_id):
    """Get a specific member by ID (multi-tenant filtered)"""
    gym_id = get_current_gym_id()
    if not gym_id:
        return jsonify({'error': 'Gym ID not found in token'}), 400
    
    member = Member.query.filter_by(id=member_id, gym_id=gym_id).first()
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    # Log the view operation
    ActivityLogger.log_view('member', entity_id=member_id, gym_id=gym_id)
    
    return jsonify({'member': member.to_dict()}), 200


@members_bp.route('/<int:member_id>', methods=['PUT'])
@jwt_required()
def update_member(member_id):
    """Update a member (multi-tenant filtered)"""
    gym_id = get_current_gym_id()
    if not gym_id:
        return jsonify({'error': 'Gym ID not found in token'}), 400
    
    member = Member.query.filter_by(id=member_id, gym_id=gym_id).first()
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    data = request.get_json() or {}
    
    try:
        # Parse dates if provided
        if data.get('date_of_birth'):
            member.date_of_birth = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()
        if data.get('membership_start_date'):
            member.membership_start_date = datetime.strptime(data['membership_start_date'], '%Y-%m-%d').date()
        if data.get('membership_end_date'):
            member.membership_end_date = datetime.strptime(data['membership_end_date'], '%Y-%m-%d').date()
        
        # Update allowed fields
        if 'first_name' in data:
            member.first_name = data['first_name']
        if 'last_name' in data:
            member.last_name = data['last_name']
        if 'gender' in data:
            member.gender = data['gender']
        if 'phone' in data:
            # Check unique constraint if phone is being changed
            if data['phone'] != member.phone:
                existing_phone = Member.query.filter_by(gym_id=gym_id, phone=data['phone']).first()
                if existing_phone:
                    return jsonify({'error': 'Phone number already exists for a member in this gym'}), 400
            member.phone = data['phone']
        if 'email' in data:
            # Check unique constraint if email is being changed
            if data['email'] != member.email:
                existing_email = Member.query.filter_by(gym_id=gym_id, email=data['email']).first()
                if existing_email:
                    return jsonify({'error': 'Email already exists for a member in this gym'}), 400
            member.email = data['email']
        if 'address' in data:
            member.address = data['address']
        if 'emergency_contact_name' in data:
            member.emergency_contact_name = data['emergency_contact_name']
        if 'emergency_contact_phone' in data:
            member.emergency_contact_phone = data['emergency_contact_phone']
        if 'medical_notes' in data:
            member.medical_notes = data['medical_notes']
        if 'membership_plan_name' in data:
            member.membership_plan_name = data['membership_plan_name']
        if 'status' in data:
            member.status = data['status']
        if 'workout_duration_minutes' in data:
            member.workout_duration_minutes = int(data['workout_duration_minutes'])
        
        # Handle password update
        if data.get('password'):
            member.password_hash = bcrypt.generate_password_hash(data['password']).decode('utf-8')
        
        member.updated_at = datetime.utcnow()
        db.session.commit()
        
        # Log the update operation
        member_name = f"{member.first_name} {member.last_name}".strip()
        ActivityLogger.log_update('member', member_id, entity_name=member_name, gym_id=gym_id)
        
        return jsonify({
            'message': 'Member updated successfully',
            'member': member.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to update member: {str(e)}'}), 500


@members_bp.route('/<int:member_id>', methods=['DELETE'])
@jwt_required()
def delete_member(member_id):
    """Delete a member (multi-tenant filtered) - only allowed if no related records exist"""
    from app.models import Attendance, Payment
    
    gym_id = get_current_gym_id()
    if not gym_id:
        return jsonify({'error': 'Gym ID not found in token'}), 400
    
    member = Member.query.filter_by(id=member_id, gym_id=gym_id).first()
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    try:
        # Check for related records
        attendance_count = Attendance.query.filter_by(member_id=member_id, gym_id=gym_id).count()
        payment_count = Payment.query.filter_by(member_id=member_id, gym_id=gym_id).count()
        
        if attendance_count > 0 or payment_count > 0:
            return jsonify({
                'error': 'Cannot delete member with historical records',
                'message': f'This member has {attendance_count} attendance records and {payment_count} payment records. Please deactivate the member instead to preserve historical data.'
            }), 400
        
        member_name = f"{member.first_name} {member.last_name}".strip()
        
        db.session.delete(member)
        db.session.commit()
        
        # Log the delete operation
        ActivityLogger.log_delete('member', member_id, entity_name=member_name, gym_id=gym_id)
        
        return jsonify({'message': 'Member deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        # Return user-friendly error message
        return jsonify({
            'error': 'Failed to delete member',
            'message': 'This member cannot be deleted. Please deactivate the member instead to preserve historical data.'
        }), 400


@members_bp.route('/<int:member_id>/deactivate', methods=['POST'])
@jwt_required()
def deactivate_member(member_id):
    """Deactivate a member (multi-tenant filtered) - preserves historical data"""
    gym_id = get_current_gym_id()
    if not gym_id:
        return jsonify({'error': 'Gym ID not found in token'}), 400
    
    member = Member.query.filter_by(id=member_id, gym_id=gym_id).first()
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    if member.status == 'Inactive':
        return jsonify({'error': 'Member is already inactive'}), 400
    
    try:
        member.status = 'Inactive'
        member.updated_at = datetime.utcnow()
        db.session.commit()
        
        # Log the deactivate operation (don't fail if logging fails)
        try:
            member_name = f"{member.first_name} {member.last_name}".strip()
            ActivityLogger.log_update(
                'member', 
                member_id, 
                entity_name=member_name, 
                gym_id=gym_id,
                changes={'status': 'Active -> Inactive'}
            )
        except Exception as log_error:
            # Log failure but don't fail the operation
            print(f"Failed to log deactivate operation: {log_error}")
        
        return jsonify({
            'message': 'Member deactivated successfully',
            'member': member.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to deactivate member: {str(e)}'}), 500


@members_bp.route('/<int:member_id>/reactivate', methods=['POST'])
@jwt_required()
def reactivate_member(member_id):
    """Reactivate a previously deactivated member (multi-tenant filtered)"""
    gym_id = get_current_gym_id()
    if not gym_id:
        return jsonify({'error': 'Gym ID not found in token'}), 400
    
    member = Member.query.filter_by(id=member_id, gym_id=gym_id).first()
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    if member.status != 'Inactive':
        return jsonify({'error': 'Member is not inactive'}), 400
    
    try:
        member.status = 'Active'
        member.updated_at = datetime.utcnow()
        db.session.commit()
        
        # Log the reactivate operation (don't fail if logging fails)
        try:
            member_name = f"{member.first_name} {member.last_name}".strip()
            ActivityLogger.log_update(
                'member', 
                member_id, 
                entity_name=member_name, 
                gym_id=gym_id,
                changes={'status': 'Inactive -> Active'}
            )
        except Exception as log_error:
            # Log failure but don't fail the operation
            print(f"Failed to log reactivate operation: {log_error}")
        
        return jsonify({
            'message': 'Member reactivated successfully',
            'member': member.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to reactivate member: {str(e)}'}), 500


@members_bp.route('/search', methods=['GET'])
@jwt_required()
def search_members():
    """Search members by name, email, phone, or member_id"""
    gym_id = get_current_gym_id()
    if not gym_id:
        return jsonify({'error': 'Gym ID not found in token'}), 400
    
    query_param = request.args.get('q', '').strip()
    if not query_param:
        return jsonify({'members': []}), 200
    
    search_pattern = f"%{query_param}%"
    members = Member.query.filter_by(gym_id=gym_id).filter(
        or_(
            Member.first_name.ilike(search_pattern),
            Member.last_name.ilike(search_pattern),
            Member.email.ilike(search_pattern),
            Member.phone.ilike(search_pattern),
            Member.member_id.ilike(search_pattern)
        )
    ).order_by(Member.first_name.asc()).limit(20).all()
    
    return jsonify({
        'members': [member.to_dict() for member in members]
    }), 200


@members_bp.route('/me', methods=['GET'])
@jwt_required()
def get_current_member():
    """Get current logged-in member's data (for member dashboard)"""
    claims = get_jwt()
    member_id = claims.get('member_id')
    gym_id = claims.get('gym_id')
    
    if not member_id or claims.get('role') != 'member':
        return jsonify({'error': 'Access denied'}), 403
    
    member = Member.query.filter_by(id=member_id, gym_id=gym_id).first()
    if not member:
        return jsonify({'error': 'Member not found'}), 404
    
    from app.models import Attendance, Payment, Gym
    from sqlalchemy import func
    from datetime import date, timedelta
    
    # Get gym info
    gym = Gym.query.get(gym_id)
    
    # Calculate attendance stats
    today = date.today()
    today_attendance = Attendance.query.filter_by(
        member_id=member_id, 
        gym_id=gym_id,
        attendance_date=today
    ).first()
    
    total_visits = Attendance.query.filter_by(member_id=member_id, gym_id=gym_id).count()
    
    # Monthly visits (current month)
    month_start = today.replace(day=1)
    monthly_visits = Attendance.query.filter(
        Attendance.member_id == member_id,
        Attendance.gym_id == gym_id,
        Attendance.attendance_date >= month_start
    ).count()
    
    # Last check-in
    last_attendance = Attendance.query.filter_by(
        member_id=member_id, 
        gym_id=gym_id
    ).order_by(Attendance.check_in_time.desc()).first()
    
    # Attendance history (last 10 records)
    attendance_history = Attendance.query.filter_by(
        member_id=member_id, 
        gym_id=gym_id
    ).order_by(Attendance.check_in_time.desc()).limit(10).all()
    
    # Payment stats
    total_paid = db.session.query(func.sum(Payment.payment_amount)).filter(
        Payment.member_id == member_id,
        Payment.gym_id == gym_id,
        Payment.payment_status == 'Paid'
    ).scalar() or 0
    
    pending_amount = db.session.query(func.sum(Payment.payment_amount)).filter(
        Payment.member_id == member_id,
        Payment.gym_id == gym_id,
        Payment.payment_status == 'Pending'
    ).scalar() or 0
    
    # Last payment
    last_payment = Payment.query.filter_by(
        member_id=member_id,
        gym_id=gym_id,
        payment_status='Paid'
    ).order_by(Payment.payment_date.desc()).first()
    
    # Payment history (last 10 records)
    payment_history = Payment.query.filter_by(
        member_id=member_id,
        gym_id=gym_id
    ).order_by(Payment.payment_date.desc()).limit(10).all()
    
    # Calculate remaining membership days
    remaining_days = None
    if member.membership_end_date:
        remaining_days = (member.membership_end_date - today).days
        if remaining_days < 0:
            remaining_days = 0
    
    return jsonify({
        'member': member.to_dict(),
        'gym': gym.to_dict() if gym else None,
        'attendance': {
            'today_status': today_attendance.status if today_attendance else 'Not Checked In',
            'total_visits': total_visits,
            'monthly_visits': monthly_visits,
            'last_check_in': last_attendance.check_in_time.isoformat() if last_attendance else None,
            'history': [att.to_dict() for att in attendance_history]
        },
        'payments': {
            'total_paid': float(total_paid),
            'pending_amount': float(pending_amount),
            'last_payment': last_payment.to_dict() if last_payment else None,
            'history': [pay.to_dict() for pay in payment_history]
        },
        'membership': {
            'remaining_days': remaining_days,
        }
    }), 200

@members_bp.route('/bulk-upload/template', methods=['GET'])
@jwt_required()
def download_bulk_upload_template():
    """Download Excel template for bulk member upload"""
    gym_id = get_current_gym_id()
    if not gym_id:
        return jsonify({'error': 'Gym ID not found in token'}), 400
    
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Member Import Template"
        
        # Define headers matching the Member model fields
        headers = [
            'First Name', 'Last Name', 'Gender', 'Date of Birth', 
            'Phone Number', 'Email', 'Address', 'Password',
            'Emergency Contact Name', 'Emergency Contact Phone',
            'Membership Plan Name', 'Status', 'Start Date', 'End Date', 'Workout Duration', 'Medical Notes'
        ]
        
        # Write headers with bold font
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Add example data
        example_data = [
            ['John', 'Doe', 'Male', '01-01-2000', '9876543210', 'john@gmail.com', '123 Main St', 'john123', '', '', 'Gold Plan', 'Active', '10-07-2026', '10-07-2027', '2 Hours', ''],
            ['Alex', 'Kumar', 'Male', '02-02-2000', '9876543211', 'alex@gmail.com', '456 Oak Ave', 'alex123', 'Jane Doe', '9876543212', 'Silver Plan', 'Active', '10-07-2026', '10-07-2027', '1 Hour 30 Minutes', 'No allergies']
        ]
        
        for row_num, row_data in enumerate(example_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='member_import_template.xlsx'
        )
    except Exception as e:
        return jsonify({'error': f'Failed to generate template: {str(e)}'}), 500

@members_bp.route('/bulk-upload', methods=['POST'])
@jwt_required()
def bulk_upload_members():
    """Bulk upload members from Excel file"""
    gym_id = get_current_gym_id()
    if not gym_id:
        return jsonify({'error': 'Gym ID not found in token'}), 400
    
    # Check if file is present
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    # Validate file extension
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Invalid file type. Please upload .xlsx or .xls file'}), 400
    
    try:
        # Read Excel file
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Get all rows
        rows = list(ws.iter_rows(values_only=True))
        
        # Validate header row
        expected_headers = [
            'First Name', 'Last Name', 'Gender', 'Date of Birth', 
            'Phone Number', 'Email', 'Address', 'Password',
            'Emergency Contact Name', 'Emergency Contact Phone',
            'Membership Plan Name', 'Status', 'Start Date', 'End Date', 'Workout Duration', 'Medical Notes'
        ]
        
        if not rows or len(rows) < 1:
            return jsonify({'error': 'File is empty'}), 400
        
        actual_headers = [str(h).strip() for h in rows[0]]
        if actual_headers != expected_headers:
            return jsonify({
                'error': 'Invalid template format. Please download the correct template.',
                'expected': expected_headers,
                'actual': actual_headers
            }), 400
        
        # Process data rows (skip header)
        data_rows = rows[1:]
        
        # Validate max rows
        if len(data_rows) > 500:
            return jsonify({'error': 'Maximum 500 members allowed per upload'}), 400
        
        # Get existing emails and phones for validation
        existing_members = Member.query.filter_by(gym_id=gym_id).all()
        existing_emails = {m.email.lower() for m in existing_members if m.email}
        existing_phones = {m.phone for m in existing_members if m.phone}
        
        # Get existing membership plans for validation
        membership_plans = MembershipPlan.query.filter_by(gym_id=gym_id).all()
        existing_plan_names = {plan.plan_name.lower(): plan.plan_name for plan in membership_plans}
        
        print(f"DEBUG: Gym ID: {gym_id}")
        print(f"DEBUG: Existing membership plans: {existing_plan_names}")
        print(f"DEBUG: Total rows in file: {len(data_rows)}")
        
        # Track validation errors
        errors = []
        valid_members = []
        
        # Track emails and phones in current upload
        upload_emails = {}
        upload_phones = {}
        
        for row_num, row in enumerate(data_rows, start=2):  # Start from row 2 (after header)
            if not row or all(cell is None for cell in row):
                print(f"DEBUG: Row {row_num} is empty, skipping")
                continue  # Skip empty rows
            
            try:
                first_name = str(row[0]).strip() if row[0] else ''
                last_name = str(row[1]).strip() if row[1] else ''
                gender = str(row[2]).strip() if row[2] else ''
                dob = str(row[3]).strip() if row[3] else ''
                phone = str(row[4]).strip() if row[4] else ''
                email = str(row[5]).strip().lower() if row[5] else ''
                address = str(row[6]).strip() if row[6] else ''
                password = str(row[7]).strip() if row[7] else ''
                emergency_contact_name = str(row[8]).strip() if row[8] else ''
                emergency_contact_phone = str(row[9]).strip() if row[9] else ''
                membership_plan_name = str(row[10]).strip() if row[10] else ''
                status = str(row[11]).strip() if row[11] else ''
                start_date = str(row[12]).strip() if row[12] else ''
                end_date = str(row[13]).strip() if row[13] else ''
                workout_duration = str(row[14]).strip() if row[14] else ''
                medical_notes = str(row[15]).strip() if row[15] else ''
                
                print(f"DEBUG: Processing row {row_num}: {first_name} {last_name}, Plan: {membership_plan_name}")
                print(f"DEBUG: Row {row_num} data - Phone: {phone}, Email: {email}, Password: {'*' * len(password) if password else 'None'}")
                
                # Validation
                if not first_name:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'First Name is required'})
                    print(f"DEBUG: Row {row_num} failed - First Name required")
                    continue
                
                if not phone:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Phone Number is required'})
                    continue
                
                # Phone format validation (min 10 digits)
                phone_digits = ''.join(filter(str.isdigit, phone))
                if len(phone_digits) < 10:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Phone number must have at least 10 digits'})
                    continue
                
                # Check duplicate phone within upload
                if phone in upload_phones:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Duplicate phone within upload'})
                    continue
                upload_phones[phone] = row_num
                
                # Check existing phone in database
                if phone in existing_phones:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Phone already exists in database'})
                    continue
                
                if not email:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Email is required'})
                    continue
                
                # Email format validation
                if '@' not in email or '.' not in email:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Invalid email format'})
                    continue
                
                # Check duplicate email within upload
                if email in upload_emails:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Duplicate email within upload'})
                    continue
                upload_emails[email] = row_num
                
                # Check existing email in database
                if email in existing_emails:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Email already exists in database'})
                    continue
                
                if not password:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Password is required'})
                    continue
                
                if not membership_plan_name:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Membership Plan Name is required'})
                    continue
                
                # Validate membership plan exists in database
                if membership_plan_name.lower() not in existing_plan_names:
                    errors.append({
                        'row': row_num, 
                        'name': first_name, 
                        'error': f'Membership Plan "{membership_plan_name}" not found. Available plans: {", ".join(existing_plan_names.values())}'
                    })
                    continue
                
                if not status:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Status is required'})
                    continue
                
                if status not in ['Active', 'Inactive', 'Expired']:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Invalid Status. Must be Active, Inactive, or Expired'})
                    continue
                
                if not start_date:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Start Date is required'})
                    continue
                
                if not end_date:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'End Date is required'})
                    continue
                
                # Validate workout duration
                workout_duration_mapping = {
                    '30 minutes': 30,
                    '1 hour': 60,
                    '1 hour 30 minutes': 90,
                    '2 hours': 120,
                    '2 hours 30 minutes': 150
                }
                
                if not workout_duration:
                    errors.append({'row': row_num, 'name': first_name, 'error': 'Workout Duration is required'})
                    continue
                
                workout_duration_lower = workout_duration.lower().strip()
                if workout_duration_lower not in workout_duration_mapping:
                    errors.append({
                        'row': row_num, 
                        'name': first_name, 
                        'error': f'Invalid Workout Duration. Must be one of: {", ".join(workout_duration_mapping.keys())}'
                    })
                    continue
                
                workout_duration_minutes = workout_duration_mapping[workout_duration_lower]
                
                # Validate date formats
                try:
                    parsed_start_date = datetime.strptime(start_date, '%d-%m-%Y').date()
                except ValueError:
                    try:
                        parsed_start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                    except ValueError:
                        errors.append({'row': row_num, 'name': first_name, 'error': 'Invalid Start Date format. Use DD-MM-YYYY or YYYY-MM-DD'})
                        continue
                
                try:
                    parsed_end_date = datetime.strptime(end_date, '%d-%m-%Y').date()
                except ValueError:
                    try:
                        parsed_end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                    except ValueError:
                        errors.append({'row': row_num, 'name': first_name, 'error': 'Invalid End Date format. Use DD-MM-YYYY or YYYY-MM-DD'})
                        continue
                
                if dob:
                    try:
                        parsed_dob = datetime.strptime(dob, '%d-%m-%Y').date()
                    except ValueError:
                        try:
                            parsed_dob = datetime.strptime(dob, '%Y-%m-%d').date()
                        except ValueError:
                            errors.append({'row': row_num, 'name': first_name, 'error': 'Invalid Date of Birth format. Use DD-MM-YYYY or YYYY-MM-DD'})
                            continue
                else:
                    parsed_dob = None
                
                # Add to valid members
                valid_members.append({
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email,
                    'phone': phone,
                    'gender': gender,
                    'date_of_birth': parsed_dob,
                    'address': address,
                    'password': password,
                    'emergency_contact_name': emergency_contact_name,
                    'emergency_contact_phone': emergency_contact_phone,
                    'membership_plan_name': membership_plan_name,
                    'status': status,
                    'membership_start_date': parsed_start_date,
                    'membership_end_date': parsed_end_date,
                    'workout_duration_minutes': workout_duration_minutes,
                    'medical_notes': medical_notes
                })
                
            except Exception as e:
                print(f"DEBUG: Exception processing row {row_num}: {str(e)}")
                import traceback
                traceback.print_exc()
                errors.append({'row': row_num, 'name': str(row[0]) if row else 'Unknown', 'error': f'Processing error: {str(e)}'})
                continue
        
        print(f"DEBUG: Validation complete. Errors: {len(errors)}, Valid members: {len(valid_members)}")
        
        # If there are errors, return them without importing
        if errors:
            print(f"DEBUG: Returning validation errors: {errors}")
            return jsonify({
                'error': 'Validation failed',
                'total_records': len(data_rows),
                'success_count': 0,
                'failed_count': len(errors),
                'errors': errors
            }), 400
        
        # Import valid members using the same logic as create_member
        imported_count = 0
        import_errors = []
        
        print(f"DEBUG: Starting import of {len(valid_members)} valid members")
        
        for idx, member_data in enumerate(valid_members):
            try:
                print(f"DEBUG: Importing member {idx + 1}/{len(valid_members)}: {member_data['first_name']} {member_data['last_name']}")
                
                # Generate unique member_id
                import uuid
                member_id = f"MEM{str(uuid.uuid4())[:8].upper()}"
                
                # Ensure member_id is unique within gym
                while Member.query.filter_by(gym_id=gym_id, member_id=member_id).first():
                    member_id = f"MEM{str(uuid.uuid4())[:8].upper()}"
                
                print(f"DEBUG: Generated member_id: {member_id}")
                
                # Hash password
                password_hash = bcrypt.generate_password_hash(member_data['password']).decode('utf-8')
                print(f"DEBUG: Password hashed successfully")
                
                # Create member using the same logic as create_member
                member = Member(
                    gym_id=gym_id,
                    member_id=member_id,
                    first_name=member_data['first_name'],
                    last_name=member_data['last_name'],
                    gender=member_data['gender'],
                    date_of_birth=member_data['date_of_birth'],
                    phone=member_data['phone'],
                    email=member_data['email'],
                    password_hash=password_hash,
                    address=member_data['address'],
                    emergency_contact_name=member_data['emergency_contact_name'],
                    emergency_contact_phone=member_data['emergency_contact_phone'],
                    medical_notes=member_data['medical_notes'],
                    membership_plan_name=member_data['membership_plan_name'],
                    membership_start_date=member_data['membership_start_date'],
                    membership_end_date=member_data['membership_end_date'],
                    status=member_data['status'],
                    workout_duration_minutes=member_data['workout_duration_minutes']
                )
                
                db.session.add(member)
                imported_count += 1
                print(f"DEBUG: Member added to session. Total imported: {imported_count}")
                
            except Exception as e:
                print(f"DEBUG: Exception importing member {member_data['first_name']} {member_data['last_name']}: {str(e)}")
                import traceback
                traceback.print_exc()
                import_errors.append({
                    'name': f"{member_data['first_name']} {member_data['last_name']}",
                    'error': str(e)
                })
                db.session.rollback()
                continue
        
        # Commit all successful imports
        if imported_count > 0:
            try:
                print(f"DEBUG: Committing {imported_count} members to database")
                db.session.commit()
                ActivityLogger.log_create('member', gym_id=gym_id, details={'bulk_import': True, 'count': imported_count})
                print(f"DEBUG: Commit successful")
            except Exception as e:
                print(f"DEBUG: Exception during commit: {str(e)}")
                import traceback
                traceback.print_exc()
                db.session.rollback()
                return jsonify({
                    'error': f'Failed to save members: {str(e)}',
                    'total_records': len(data_rows),
                    'success_count': 0,
                    'failed_count': len(data_rows)
                }), 500
        
        # Generate error report if there were import errors
        error_report_url = None
        if import_errors:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['Member Name', 'Error'])
            for error in import_errors:
                writer.writerow([error['name'], error['error']])
            
            # Save error report to a temporary file or return as base64
            # For simplicity, we'll return the errors in the response
            error_report_url = '/api/members/bulk-upload/error-report'
        
        return jsonify({
            'message': 'Bulk import completed',
            'total_records': len(data_rows),
            'success_count': imported_count,
            'failed_count': len(import_errors),
            'errors': import_errors if import_errors else None,
            'error_report_url': error_report_url if import_errors else None
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process file: {str(e)}'}), 500
