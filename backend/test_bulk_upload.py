import openpyxl
import requests
import json

# Test the bulk upload feature
BASE_URL = 'http://127.0.0.1:5000'

# First, we need to get a JWT token by logging in
# This requires a valid gym owner account
# For testing, we'll create a test Excel file and verify the structure

def create_test_excel():
    """Create a test Excel file with the correct structure"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Members"
    
    # Headers matching the backend
    headers = [
        'First Name', 'Last Name', 'Gender', 'Date of Birth', 
        'Phone Number', 'Email', 'Address', 'Password',
        'Emergency Contact Name', 'Emergency Contact Phone',
        'Membership Plan Name', 'Status', 'Start Date', 'End Date', 'Medical Notes'
    ]
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Add test data
    test_data = [
        ['Test', 'User1', 'Male', '01-01-1990', '9999999991', 'test1@example.com', '123 Test St', 'password123', '', '', 'Test Plan', 'Active', '10-07-2026', '10-07-2027', ''],
        ['Test', 'User2', 'Female', '02-02-1991', '9999999992', 'test2@example.com', '456 Test Ave', 'password123', '', '', 'Test Plan', 'Active', '10-07-2026', '10-07-2027', ''],
        ['Test', 'User3', 'Male', '03-03-1992', '9999999993', 'test3@example.com', '789 Test Blvd', 'password123', '', '', 'Test Plan', 'Active', '10-07-2026', '10-07-2027', '']
    ]
    
    for row_num, row_data in enumerate(test_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    wb.save('test_bulk_upload.xlsx')
    print("Test Excel file created: test_bulk_upload.xlsx")
    print(f"Headers: {headers}")
    print(f"Number of test rows: {len(test_data)}")

if __name__ == '__main__':
    create_test_excel()
